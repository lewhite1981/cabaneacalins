#!/usr/bin/env python3
"""
publier.py — Script de publication automatique du blog Cabane à Câlin
Lit le calendrier-blog.xlsx, et pour chaque article dont la date est arrivée
(statut ≠ Publié), active la carte dans blog.html et met à jour le xlsx.

Usage : python3 publier.py
Planifié chaque matin à 8h via le scheduler Cowork.
"""

import os
import re
import sys
from datetime import date, datetime

# ── Dépendances ──────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("[publier] Installation de openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl

# ── Chemins ──────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH  = os.path.join(BASE_DIR, "calendrier-blog.xlsx")
BLOG_HTML  = os.path.join(BASE_DIR, "blog.html")
LOG_PATH   = os.path.join(BASE_DIR, "publier.log")

def log(msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    line = f"[{timestamp}] {msg}"
    print(line)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")

# ── Lecture du calendrier ─────────────────────────────────────────────────────
def lire_calendrier():
    if not os.path.exists(XLSX_PATH):
        log(f"ERREUR : fichier introuvable → {XLSX_PATH}")
        return []

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    articles = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Colonnes : 0=Date, 1=Thème, 2=Catégorie, 3=Durée, 4=Statut, 5=Lien
        pub_date = row[0]
        theme    = row[1]
        statut   = row[4] if len(row) > 4 else ""
        if pub_date is None or theme is None:
            continue
        if isinstance(pub_date, datetime):
            pub_date = pub_date.date()
        elif isinstance(pub_date, str):
            try:
                pub_date = datetime.strptime(pub_date, "%Y-%m-%d").date()
            except ValueError:
                continue
        elif not isinstance(pub_date, date):
            continue
        articles.append({
            "row": i,
            "date": pub_date,
            "theme": theme.strip(),
            "statut": (statut or "").strip(),
        })
    return articles

# ── Déterminer le numéro d'article depuis blog.html ─────────────────────────
def trouver_numero_article(theme_recherche):
    """
    Cherche dans blog.html la carte dont le titre correspond au thème,
    et retourne le data-article="N" correspondant.
    Utilise une correspondance approximative (premiers mots du titre).
    """
    if not os.path.exists(BLOG_HTML):
        return None

    with open(BLOG_HTML, "r", encoding="utf-8") as f:
        contenu = f.read()

    # Cherche toutes les cartes avec data-article
    cartes = re.findall(
        r'data-article="(\d+)"\s+data-date="([\d-]+)"\s+data-file="([^"]+)"',
        contenu
    )
    return cartes  # [(num, date_iso, fichier), ...]

# ── Activation d'une carte dans blog.html ────────────────────────────────────
def activer_carte(num_article: str, fichier: str):
    """
    Dans blog.html, pour l'article data-article="N" :
    - Remplace le badge "À venir" par "● Publié"
    - Remplace la classe blog-badge--soon par blog-badge--live
    - Remplace blog-card (sans --published) par blog-card blog-card--published
    - Remplace le <span> titre par un <a href="...">
    - Remplace le <span class="blog-card-cta--disabled"> par un <a class="blog-card-cta">
    """
    with open(BLOG_HTML, "r", encoding="utf-8") as f:
        html = f.read()

    # On repère le bloc de l'article par data-article="N"
    pattern = (
        r'(<article\b[^>]*\bdata-article="' + re.escape(num_article) + r'"[^>]*>)'
        r'(.*?)'
        r'(</article>)'
    )
    match = re.search(pattern, html, re.DOTALL)
    if not match:
        log(f"  ⚠ Article {num_article} : bloc <article> introuvable dans blog.html")
        return False

    bloc_original = match.group(0)
    bloc = bloc_original

    # 1. Badge
    bloc = re.sub(
        r'<span class="blog-badge blog-badge--soon">.*?</span>',
        '<span class="blog-badge blog-badge--live">&#9679; Publi&eacute;</span>',
        bloc, flags=re.DOTALL
    )

    # 2. Classe article (ajouter --published)
    bloc = re.sub(
        r'class="blog-card reveal"',
        'class="blog-card reveal blog-card--published"',
        bloc
    )

    # 3. Titre : <span>Titre</span> → <a href="fichier.html">Titre</a>
    bloc = re.sub(
        r'<h2 class="blog-card-title"><span>(.*?)</span></h2>',
        lambda m: f'<h2 class="blog-card-title"><a href="{fichier}">{m.group(1)}</a></h2>',
        bloc, flags=re.DOTALL
    )

    # 4. CTA : <span class="blog-card-cta blog-card-cta--disabled">...</span>
    #       → <a href="fichier.html" class="blog-card-cta">Lire l'article →</a>
    bloc = re.sub(
        r'<span class="blog-card-cta blog-card-cta--disabled">.*?</span>',
        f'<a href="{fichier}" class="blog-card-cta">Lire l\'article &rarr;</a>',
        bloc, flags=re.DOTALL
    )

    html = html.replace(bloc_original, bloc)

    with open(BLOG_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    log(f"  ✓ blog.html mis à jour : article {num_article} ({fichier}) → Publié")
    return True

# ── Mise à jour du statut dans le xlsx ────────────────────────────────────────
def marquer_publie(row_index: int):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Colonne E (index 5) = Statut
    ws.cell(row=row_index, column=5, value="Publié")

    # Style vert pour la cellule statut
    from openpyxl.styles import PatternFill, Font
    vert = PatternFill(start_color="7A9E7E", end_color="7A9E7E", fill_type="solid")
    blanc = Font(color="FFFFFF", bold=True)
    ws.cell(row=row_index, column=5).fill = vert
    ws.cell(row=row_index, column=5).font = blanc

    wb.save(XLSX_PATH)
    log(f"  ✓ calendrier-blog.xlsx ligne {row_index} → Publié")

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    today = date.today()
    log(f"=== publier.py démarré — {today} ===")

    articles = lire_calendrier()
    if not articles:
        log("Aucun article trouvé dans le calendrier.")
        return

    cartes = trouver_numero_article("")  # retourne toutes les cartes
    # cartes = [(num_str, date_iso, fichier), ...]
    carte_par_date = {c[1]: (c[0], c[2]) for c in cartes}

    publies = 0
    for art in articles:
        if art["statut"].lower() in ("publié", "publie", "publié"):
            continue  # déjà publié

        if art["date"] <= today:
            date_iso = art["date"].strftime("%Y-%m-%d")
            if date_iso in carte_par_date:
                num, fichier = carte_par_date[date_iso]
                log(f"→ Publication : « {art['theme']} » ({date_iso})")
                ok = activer_carte(num, fichier)
                if ok:
                    marquer_publie(art["row"])
                    publies += 1
            else:
                log(f"  ⚠ Pas de carte trouvée pour date={date_iso} dans blog.html")

    if publies == 0:
        log("Aucun nouvel article à publier aujourd'hui.")
    else:
        log(f"=== {publies} article(s) publié(s) avec succès ===")

if __name__ == "__main__":
    main()
